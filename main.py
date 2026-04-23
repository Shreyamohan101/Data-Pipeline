import csv
import psycopg2
from openpyxl import load_workbook
from dotenv import load_dotenv
import os
import json
import pandas as pd
from sqlalchemy import create_engine

load_dotenv(r"C:\Users\ASUS\Desktop\imei-pipeline\.env")

password = os.getenv("password")

if not password:
    raise Exception("Password not found in .env")

excel_file = r"C:\IMPORTSERVER\IMEI\IMEI.xlsx"
csv_file = r"C:\IMPORTSERVER\IMEI\temp.csv"

db_config = {
    "host": "localhost",
    "database": "imei_db",
    "user": "postgres",
    "password": password,
    "port": 5432
}

conn = psycopg2.connect(**db_config)
cur = conn.cursor()

# SQLAlchemy engine (for staging)
from urllib.parse import quote_plus

encoded_password = quote_plus(password)

engine = create_engine(
    f"postgresql://postgres:{encoded_password}@localhost:5432/imei_db"
)

# STEP 1 — LOAD IMEI TABLE

wb = load_workbook(excel_file, read_only=True, data_only=True)
ws = wb.active

rows = ws.iter_rows(values_only=True)
headers = next(rows)

def clean(col):
    return str(col).strip().replace(" ", "_").replace("-", "_").lower()

columns = [clean(col) for col in headers]

table_name = "imei_data"

cur.execute(f"DROP TABLE IF EXISTS {table_name};")

create_query = f"""
CREATE TABLE {table_name} (
    {', '.join([f'"{col}" TEXT' for col in columns])}
);
"""
cur.execute(create_query)

with open(csv_file, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(columns)

    for row in rows:
        writer.writerow([str(cell) if cell is not None else "" for cell in row])

with open(csv_file, "r", encoding="utf-8") as f:
    cur.copy_expert(f"COPY {table_name} FROM STDIN WITH CSV HEADER", f)

# STEP 2 — ADD DATE COLUMNS

cur.execute("""
ALTER TABLE imei_data
ADD COLUMN act_date TEXT,
ADD COLUMN act_day TEXT;
""")

cur.execute("""
UPDATE imei_data
SET 
    act_date = TO_CHAR(activation_time::timestamp, 'DD-MM-YYYY'),
    act_day  = TRIM(TO_CHAR(activation_time::timestamp, 'Day'))
WHERE activation_time ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}';
""")

# STEP 3 — LOAD CONFIG

with open("config.json") as f:
    config = json.load(f)

# STEP 4 — STAGING LOOP

for file in config["files"]:
    print(f"Processing {file['sheet_name']}...")

    # TEMP DEBUG (to find correct header row)
    df = pd.read_excel(
        file["file_path"],
        sheet_name=file["sheet_name"],
        header=None
    )

    print(df.head(10))  
    df = pd.read_excel(
    file["file_path"],
    sheet_name=file["sheet_name"],
    header=file.get("header", 0)
)

    df.columns = [
        str(c).strip().lower().replace(" ", "_").replace("-", "_")
        for c in df.columns
    ]

    print("Columns:", df.columns)

    df.to_sql(file["table_name"], engine, if_exists="replace", index=False)

    print(f"Loaded {file['table_name']}")

cur.execute("CREATE INDEX IF NOT EXISTS idx_imei_rtl_id ON imei_data(rtl_id);")
cur.execute("CREATE INDEX IF NOT EXISTS idx_fos_retailer_id ON stg_fos_map(retailer_id);")
cur.execute("CREATE INDEX IF NOT EXISTS idx_soms_retailer_id ON stg_soms(retailer_id);")

# STEP 5 — ENRICHMENT LOOP

for file in config["files"]:
    table = file["table_name"]
    join = file["join"]

    for col in file["columns"]:
        print(f"Updating {col} from {table}")

        cur.execute(f"""
        ALTER TABLE imei_data 
        ADD COLUMN IF NOT EXISTS {col} TEXT;
        """)

        cur.execute(f"""
        UPDATE imei_data i
        SET {col} = s.{col}
        FROM {table} s
        WHERE i.{join['imei_column']}::TEXT = s.{join['file_column']}::TEXT;
        """)

# FINAL COMMIT

conn.commit()
cur.close()
conn.close()

print("DONE — FULL PIPELINE EXECUTED")